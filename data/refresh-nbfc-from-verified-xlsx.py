#!/usr/bin/env python3
"""Refresh Big Screen NBFC JSON from VERIFIED loan-book workbook."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data/nbfc_loanbook_strictrefresh_VERIFIED_20260813.xlsx"
OUT_STATS = ROOT / "data/nbfc-country-stats.json"
OUT_STATS_WEB = ROOT / "web/src/data/nbfc-country-stats.json"
OUT_CLASS = ROOT / "data/country-classifications.json"
OUT_CLASS_WEB = ROOT / "web/src/data/country-classifications.json"

IMF_DEV = {
    "Developed": "发达",
    "Developing/Emerging": "发展中/新兴",
    "Developing": "发展中/新兴",
}
IMF_WEO = {
    "Advanced economy": "发达经济体",
    "Emerging market and developing economy": "新兴市场与发展中经济体",
}
WB_INCOME = {
    "High income": "高收入",
    "Upper middle income": "中高收入",
    "Lower middle income": "中低收入",
    "Low income": "低收入",
}
QUALITY = {
    "official": "official",
    "semi-official": "semi-official",
    "secondary": "secondary",
    "not_found": "not_found",
}

MASTER_HEADERS = [
    "country_code",
    "country_name_en",
    "country_name_zh",
    "refresh_status",
    "recommended_scope",
    "recommended_metric",
    "outstanding_local",
    "currency",
    "usd_bn_indicative",
    "as_of",
    "quality",
    "source_url",
    "evidence_file",
    "prior_json_usd_bn",
    "notes",
    "development_label_imf",
    "imf_weo_group",
    "wb_income_group",
    "wb_income_code",
    "wb_region",
    "oecd_member",
    "class_tag",
    "classification_notes",
    "institution_count",
    "institution_type",
    "institution_as_of",
    "institution_quality",
    "institution_source_url",
    "verification_comments_20260811",
    "usd_bn_current_fx_20260811",
    "gross_outstanding_incl_mortgage_local",
    "mortgage_portion_local",
    "lender_type_classification",
    "borrowers_count",
    "borrowers_quality",
    "borrowers_basis",
    "borrowers_as_of",
    "borrowers_source_url",
    "avg_ticket_size_local",
    "ticket_quality",
    "dpd_ratio",
    "dpd_quality",
    "dpd_basis",
    "dpd_as_of",
    "dpd_source_url",
]


def sheet_dicts(wb, name: str) -> list[dict]:
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def fix_master_loan_rows(wb) -> list[dict]:
    """Country_Master_Loan may have A1:B2 merge + MENA shift; realign by Chinese name."""
    ws = wb["Country_Master_Loan"]
    headers = MASTER_HEADERS
    zh_to_meta: dict[str, tuple[str, str]] = {}
    inst_order: list[str] = []
    for r in sheet_dicts(wb, "Country_Master_Institutions"):
        code = r.get("country_code")
        zh = (r.get("country_name_zh") or "").strip()
        en = (r.get("country_name_en") or "").strip()
        if code and zh:
            zh_to_meta[zh] = (code, en)
            inst_order.append(code)

    by_code: dict[str, dict] = {}
    for r_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(r_idx, c).value for c in range(1, len(headers) + 1)]
        d = dict(zip(headers, vals))
        zh = (d.get("country_name_zh") or "").strip()
        meta = zh_to_meta.get(zh)
        if not meta:
            code = d.get("country_code")
            if isinstance(code, str) and re.fullmatch(r"[A-Z]{2}", code):
                by_code[code] = d
            continue
        code, en = meta
        d["country_code"] = code
        d["country_name_en"] = en
        d["country_name_zh"] = zh
        by_code[code] = d

    return [by_code[c] for c in inst_order if c in by_code]


def fmt_count(count, type_zh: str, as_of: str) -> str:
    if count is None or count == "":
        return ""
    try:
        n = int(count)
        num = f"{n:,}"
    except Exception:
        num = str(count)
    bits = [num]
    label = (type_zh or "").strip()
    if label:
        bits.append(f"（{label}")
        if as_of:
            bits.append(f"；{as_of}）")
        else:
            bits.append("）")
        return "".join(bits)
    if as_of:
        return f"{num}（{as_of}）"
    return num


def fmt_usd(bn) -> tuple[str, float | None]:
    if bn is None or bn == "":
        return "", None
    try:
        v = float(bn)
    except Exception:
        return "", None
    if not (v > 0):
        return "", None
    if v >= 100:
        return f"约 USD {v:.1f} bn", round(v, 1)
    return f"约 USD {v:.2f} bn", round(v, 2)


def fmt_local(amount, currency: str, label: str) -> str:
    if amount is None or amount == "":
        return ""
    try:
        a = float(amount)
    except Exception:
        return f"{amount}（{label}）" if label else str(amount)
    cur = (currency or "").upper()
    lab = label or "outstanding"
    if cur == "CNY":
        return f"{a / 1e8:,.2f} 亿元人民币（{lab}）"
    if cur == "INR":
        crore = a / 1e7
        return f"₹{crore:,.0f} crore（{lab}）"
    if cur == "IDR":
        return f"Rp{a / 1e12:.2f} 万亿（{lab}）"
    if a >= 1e12:
        return f"{a / 1e12:,.2f} 万亿{cur}（{lab}）"
    if a >= 1e9:
        return f"{a / 1e9:,.2f} 十亿{cur}（{lab}）"
    if a >= 1e6:
        return f"{a / 1e6:,.2f} 百万{cur}（{lab}）"
    return f"{a:,.0f} {cur}（{lab}）"


def fmt_borrowers(count, as_of: str = "") -> str:
    if count is None or count == "":
        return ""
    try:
        n = float(count)
    except Exception:
        return str(count)
    if not (n > 0):
        return ""
    if n >= 1e8:
        s = f"{n / 1e8:.2f} 亿人"
    elif n >= 1e4:
        s = f"{n / 1e4:,.1f} 万人"
    else:
        s = f"{int(round(n)):,} 人"
    if as_of:
        return f"{s}（{as_of}）"
    return s


def fmt_avg_ticket(amount, currency: str, as_of: str = "") -> str:
    if amount is None or amount == "":
        return ""
    try:
        a = float(amount)
    except Exception:
        return str(amount)
    if not (a > 0):
        return ""
    cur = (currency or "").upper()
    if cur == "CNY":
        s = f"约 ¥{a:,.0f}"
    elif cur == "INR":
        s = f"约 ₹{a:,.0f}"
    elif cur == "IDR":
        s = f"约 Rp{a:,.0f}"
    elif cur:
        s = f"约 {a:,.0f} {cur}"
    else:
        s = f"约 {a:,.0f}"
    if as_of:
        return f"{s}（{as_of}）"
    return s


def fmt_dpd(ratio, as_of: str = "", basis: str = "") -> str:
    if ratio is None or ratio == "":
        return ""
    try:
        r = float(ratio)
    except Exception:
        return str(ratio)
    # workbook stores decimal ratio (0.029 -> 2.9%)
    pct = r * 100 if abs(r) <= 1.5 else r
    s = f"{pct:.2f}%"
    bits = [s]
    if basis:
        bits.append(str(basis).strip())
    if as_of:
        bits.append(str(as_of).strip())
    if len(bits) == 1:
        return s
    return f"{s}（{'；'.join(bits[1:])}）"


def map_quality(q) -> str:
    if not q:
        return "not_found"
    return QUALITY.get(str(q).strip(), "secondary")


def clean_as_of(raw) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    low = s.lower()
    # workbook sometimes puts methodology notes into as_of
    if "estimate" in low or "see col" in low or "见列" in s:
        return ""
    return s


def metrics_from_master(m: dict) -> tuple[str, str, str]:
    """borrowers_count / avg_ticket_size_local / dpd_ratio → JSON display fields."""
    b_asof = clean_as_of(m.get("borrowers_as_of")) or clean_as_of(m.get("as_of"))
    d_asof = clean_as_of(m.get("dpd_as_of")) or clean_as_of(m.get("as_of"))
    borrowers = fmt_borrowers(m.get("borrowers_count"), b_asof)
    avg = fmt_avg_ticket(
        m.get("avg_ticket_size_local"),
        str(m.get("currency") or ""),
        b_asof,
    )
    basis = str(m.get("dpd_basis") or "").strip()
    # keep basis short in cell
    if len(basis) > 48:
        basis = basis[:46] + "…"
    dpd = fmt_dpd(m.get("dpd_ratio"), d_asof, basis)
    return borrowers, avg, dpd


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source workbook: {SRC}")
    wb = load_workbook(SRC, data_only=True)
    master = fix_master_loan_rows(wb)
    inst_detail = sheet_dicts(wb, "Institution_Count_Detail")
    loan_detail = sheet_dicts(wb, "Loan_Book_Detail")

    inst_by_code: dict[str, list[dict]] = {}
    for r in inst_detail:
        code = r.get("country_code")
        if code:
            inst_by_code.setdefault(code, []).append(r)

    loan_by_code: dict[str, list[dict]] = {}
    for r in loan_detail:
        code = r.get("country_code")
        if code:
            loan_by_code.setdefault(code, []).append(r)

    def pick_inst(code: str, prefer_substr: str | None = None) -> dict | None:
        rows = inst_by_code.get(code, [])
        if prefer_substr:
            for r in rows:
                blob = f"{r.get('institution_type_en') or ''} {r.get('institution_type_zh') or ''}"
                if prefer_substr.lower() in blob.lower():
                    return r
        for r in rows:
            if str(r.get("recommended") or "").upper() == "Y":
                return r
        return rows[0] if rows else None

    def pick_loan(code: str, prefer_substr: str | None = None) -> dict | None:
        rows = loan_by_code.get(code, [])
        if prefer_substr:
            for r in rows:
                blob = f"{r.get('metric_label') or ''} {r.get('metric_scope_en') or ''}"
                if prefer_substr.lower() in blob.lower():
                    return r
        return rows[0] if rows else None

    stats_rows: list[dict] = []
    class_rows: list[dict] = []

    for m in master:
        code = m["country_code"]
        zh = m.get("country_name_zh") or code
        class_rows.append(
            {
                "country_code": code,
                "country_name_zh": zh,
                "imf_dev_label": IMF_DEV.get(str(m.get("development_label_imf") or "").strip(), str(m.get("development_label_imf") or "")),
                "imf_weo_group": IMF_WEO.get(str(m.get("imf_weo_group") or "").strip(), str(m.get("imf_weo_group") or "")),
                "wb_income_group": WB_INCOME.get(str(m.get("wb_income_group") or "").strip(), str(m.get("wb_income_group") or "")),
                "wb_income_code": m.get("wb_income_code") or "",
                "wb_region": m.get("wb_region") or "",
                "oecd_member": {"Yes": "是", "No": "否"}.get(str(m.get("oecd_member") or "").strip(), str(m.get("oecd_member") or "")),
            }
        )

        borrowers, avg_ticket, dpd = metrics_from_master(m)
        ver_note = m.get("verification_comments_20260811") or ""

        if code == "CN":
            pairs = [
                ("Consumer finance", "消费金融", "消费金融公司"),
                ("Microloan", "小额贷款", "小额贷款公司"),
            ]
            for loan_key, inst_key, name_zh in pairs:
                ld = pick_loan(code, loan_key)
                inst = pick_inst(code, inst_key)
                usd_s, usd_bn = fmt_usd((ld or {}).get("outstanding_usd_bn_indicative"))
                type_zh = (inst or {}).get("institution_type_zh") or name_zh
                stats_rows.append(
                    {
                        "country_code": code,
                        "country_name_zh": zh,
                        "nbfc_equivalent_name": type_zh,
                        "regulator": (inst or {}).get("source_org") or (ld or {}).get("source_org") or "",
                        "source_url": (ld or {}).get("source_url") or (inst or {}).get("source_url") or m.get("source_url") or "",
                        "source_title": (ld or {}).get("source_title") or (inst or {}).get("source_title") or "",
                        "as_of": (ld or {}).get("as_of") or (inst or {}).get("as_of") or m.get("as_of") or "",
                        "nbfc_count": fmt_count((inst or {}).get("count"), type_zh, (inst or {}).get("as_of") or ""),
                        "loan_book_total": fmt_local(
                            (ld or {}).get("outstanding_local"),
                            (ld or {}).get("currency") or "",
                            (ld or {}).get("metric_label") or name_zh,
                        ),
                        "borrowers_covered": borrowers,
                        "avg_loan_size": avg_ticket,
                        "default_rate": dpd,
                        "other_info": (inst or {}).get("notes") or "",
                        "data_quality": map_quality((ld or {}).get("quality") or (inst or {}).get("quality")),
                        "notes": f"2026-08-13 VERIFIED workbook；borrowers/avg_ticket/dpd 取自 Country_Master_Loan；{ver_note}".strip("；"),
                        "loan_book_usd": usd_s,
                        "loan_book_usd_bn": usd_bn,
                    }
                )
            continue

        if code == "ID":
            inst_mf = pick_inst(code, "Pembiayaan") or pick_inst(code, "融资")
            inst_p2p = pick_inst(code, "P2P") or pick_inst(code, "LPBBTI")
            ld_p2p = pick_loan(code, "P2P")
            type_mf = (inst_mf or {}).get("institution_type_zh") or "融资公司"
            stats_rows.append(
                {
                    "country_code": code,
                    "country_name_zh": zh,
                    "nbfc_equivalent_name": "Perusahaan Pembiayaan / Multifinance",
                    "regulator": (inst_mf or {}).get("source_org") or "OJK",
                    "source_url": (inst_mf or {}).get("source_url") or "",
                    "source_title": (inst_mf or {}).get("source_title") or "",
                    "as_of": (inst_mf or {}).get("as_of") or "",
                    "nbfc_count": fmt_count((inst_mf or {}).get("count"), type_mf, (inst_mf or {}).get("as_of") or ""),
                    "loan_book_total": "",
                    "borrowers_covered": "",
                    "avg_loan_size": "",
                    "default_rate": "",
                    "other_info": (inst_mf or {}).get("notes") or "",
                    "data_quality": map_quality((inst_mf or {}).get("quality")),
                    "notes": "2026-08-13 VERIFIED：主表推荐在贷为 P2P；本行仅机构数",
                    "loan_book_usd": "",
                    "loan_book_usd_bn": None,
                }
            )
            type_p2p = (inst_p2p or {}).get("institution_type_zh") or "P2P网络借贷"
            usd_s, usd_bn = fmt_usd((ld_p2p or m).get("outstanding_usd_bn_indicative") if ld_p2p else m.get("usd_bn_indicative"))
            stats_rows.append(
                {
                    "country_code": code,
                    "country_name_zh": zh,
                    "nbfc_equivalent_name": "LPBBTI / Fintech P2P Lending",
                    "regulator": (inst_p2p or {}).get("source_org") or (ld_p2p or {}).get("source_org") or "OJK",
                    "source_url": (ld_p2p or {}).get("source_url") or m.get("source_url") or "",
                    "source_title": (ld_p2p or {}).get("source_title") or "",
                    "as_of": (ld_p2p or {}).get("as_of") or m.get("as_of") or "",
                    "nbfc_count": fmt_count((inst_p2p or {}).get("count"), type_p2p, (inst_p2p or {}).get("as_of") or ""),
                    "loan_book_total": fmt_local(
                        (ld_p2p or {}).get("outstanding_local") or m.get("outstanding_local"),
                        (ld_p2p or {}).get("currency") or m.get("currency") or "IDR",
                        (ld_p2p or {}).get("metric_label") or m.get("recommended_metric") or "P2P outstanding",
                    ),
                    "borrowers_covered": borrowers,
                    "avg_loan_size": avg_ticket,
                    "default_rate": dpd,
                    "other_info": (inst_p2p or {}).get("notes") or "",
                    "data_quality": map_quality((ld_p2p or {}).get("quality") or m.get("quality")),
                    "notes": f"2026-08-13 VERIFIED workbook；{m.get('notes') or ''}".strip("；"),
                    "loan_book_usd": usd_s,
                    "loan_book_usd_bn": usd_bn,
                }
            )
            continue

        inst = pick_inst(code)
        ld = None
        metric = str(m.get("recommended_metric") or "")
        if metric:
            ld = pick_loan(code, metric.split("—")[0].strip()[:24])
        if ld is None:
            for cand in loan_by_code.get(code, []):
                if cand.get("metric_label") == m.get("recommended_metric"):
                    ld = cand
                    break
        if ld is None:
            ld = pick_loan(code)

        pending = str(m.get("refresh_status") or "").startswith("PENDING")
        usd_s, usd_bn = ("", None)
        loan_total = ""
        if not pending:
            usd_s, usd_bn = fmt_usd(m.get("usd_bn_indicative"))
            loan_total = fmt_local(
                m.get("outstanding_local"),
                m.get("currency") or "",
                m.get("recommended_metric") or (ld or {}).get("metric_label") or "outstanding",
            )

        type_zh = (inst or {}).get("institution_type_zh") or ""
        type_en = (inst or {}).get("institution_type_en") or m.get("institution_type") or ""
        if type_zh:
            equiv = type_zh
        else:
            equiv = type_en or m.get("recommended_metric") or m.get("recommended_scope") or "NBFC/等效"

        stats_rows.append(
            {
                "country_code": code,
                "country_name_zh": zh,
                "nbfc_equivalent_name": equiv,
                "regulator": (inst or {}).get("source_org") or (ld or {}).get("source_org") or "",
                "source_url": (ld or {}).get("source_url") or m.get("source_url") or (inst or {}).get("source_url") or "",
                "source_title": (ld or {}).get("source_title") or (inst or {}).get("source_title") or "",
                "as_of": ("" if pending else (m.get("as_of") or (ld or {}).get("as_of") or "")) or (inst or {}).get("as_of") or "",
                "nbfc_count": fmt_count(
                    (inst or {}).get("count") if inst else m.get("institution_count"),
                    (inst or {}).get("institution_type_zh") or m.get("institution_type") or "",
                    (inst or {}).get("as_of") or m.get("institution_as_of") or "",
                ),
                "loan_book_total": loan_total,
                "borrowers_covered": "" if pending else borrowers,
                "avg_loan_size": "" if pending else avg_ticket,
                "default_rate": "" if pending else dpd,
                "other_info": (inst or {}).get("notes") or "",
                "data_quality": "not_found" if pending else map_quality(m.get("quality") or (ld or {}).get("quality")),
                "notes": (
                    f"2026-08-13 VERIFIED workbook；status={m.get('refresh_status')}"
                    + (f"；{ver_note}" if ver_note else "")
                    + (f"；{m.get('notes')}" if m.get("notes") else "")
                ),
                "loan_book_usd": usd_s,
                "loan_book_usd_bn": usd_bn,
            }
        )

    stats = {
        "meta": {
            "title": "CRM覆盖国家 · NBFC/等效非银信贷机构统计",
            "updated": "2026-08-13",
            "note": (
                "口径各国不同：印度为真NBFC；他国为最接近的非银放贷类别。数字均标注信源；未核验处留空，勿当作审计口径。"
                " 2026-08-13 按 VERIFIED 工作簿（nbfc_loanbook_strictrefresh_VERIFIED_20260813.xlsx）刷新："
                "在贷优先 Country_Master_Loan.usd_bn_indicative；机构数优先 Institution_Count_Detail（recommended=Y）；"
                "覆盖人数←borrowers_count；平均放贷额←avg_ticket_size_local；Default/NPL←dpd_ratio。"
                "中国消金+小贷分列；印尼 multifinance 仅机构数、P2P 带在贷与三项指标；热力图为国家各行之和。"
            ),
            "fx_note": (
                "放贷总量(USD)取自 VERIFIED 工作簿指示性折算（统计时点附近汇率）；"
                "仅供横向对比，不构成审计口径。"
            ),
        },
        "rows": stats_rows,
    }
    classes = {
        "meta": {
            "title": "CRM覆盖国家 · IMF/世行分类",
            "updated": "2026-08-13",
            "source": "nbfc_loanbook_strictrefresh_VERIFIED_20260813.xlsx · Country_Master_Loan",
            "note": "IMF发展标签：发达 / 发展中·新兴；世行收入：高/中高/中低/低收入。台湾等可能无世行收入码。",
        },
        "rows": class_rows,
    }

    text = json.dumps(stats, ensure_ascii=False, indent=2) + "\n"
    OUT_STATS.parent.mkdir(parents=True, exist_ok=True)
    OUT_STATS.write_text(text, encoding="utf-8")
    if OUT_STATS_WEB.parent.exists():
        OUT_STATS_WEB.write_text(text, encoding="utf-8")
    class_text = json.dumps(classes, ensure_ascii=False, indent=2) + "\n"
    OUT_CLASS.write_text(class_text, encoding="utf-8")
    if OUT_CLASS_WEB.parent.exists():
        OUT_CLASS_WEB.write_text(class_text, encoding="utf-8")

    lending = {}
    with_b = with_a = with_d = 0
    for r in stats_rows:
        bn = r.get("loan_book_usd_bn")
        if bn:
            lending[r["country_code"]] = lending.get(r["country_code"], 0) + bn
        if r.get("borrowers_covered"):
            with_b += 1
        if r.get("avg_loan_size"):
            with_a += 1
        if r.get("default_rate"):
            with_d += 1
    print(f"Wrote {OUT_STATS} rows={len(stats_rows)}")
    print(f"Countries with loan>0: {len(lending)}  sum_usd_bn={sum(lending.values()):.2f}")
    print(f"Filled borrowers={with_b} avg_ticket={with_a} dpd={with_d}")
    for c in ["CN", "ID", "IN", "TH", "PH"]:
        rows = [r for r in stats_rows if r["country_code"] == c]
        for r in rows:
            print(
                c,
                r["nbfc_equivalent_name"][:28],
                "loan",
                r.get("loan_book_usd_bn"),
                "|",
                r.get("borrowers_covered"),
                "|",
                r.get("avg_loan_size"),
                "|",
                r.get("default_rate"),
            )


if __name__ == "__main__":
    main()
