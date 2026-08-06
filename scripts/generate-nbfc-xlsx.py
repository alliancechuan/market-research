#!/usr/bin/env python3
"""Generate NBFC country stats Excel from JSON source of truth."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "web/src/data/nbfc-country-stats.json"
OUT_PATHS = [
    ROOT / "web/public/downloads/nbfc-country-stats.xlsx",
    ROOT / "web/src/assets/nbfc-country-stats.xlsx",
]

HEADERS = [
    ("country_code", "国家代码"),
    ("country_name_zh", "国家/地区"),
    ("nbfc_equivalent_name", "NBFC/等效类别"),
    ("regulator", "监管机构"),
    ("nbfc_count", "机构数量"),
    ("loan_book_total", "放贷总量"),
    ("loan_book_usd", "放贷总量(USD)"),
    ("borrowers_covered", "覆盖人数"),
    ("avg_loan_size", "平均放贷额"),
    ("default_rate", "Default/NPL"),
    ("as_of", "统计时点"),
    ("source_title", "信源标题"),
    ("source_url", "原始网站链接"),
    ("other_info", "其他信息"),
    ("data_quality", "数据质量"),
    ("notes", "备注"),
]


def main() -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    rows = data["rows"]
    meta = data.get("meta", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "NBFC统计"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E76AB")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, (_, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin

    quality_fill = {
        "official": PatternFill("solid", fgColor="E8F5E9"),
        "semi-official": PatternFill("solid", fgColor="FFF8E1"),
        "not_found": PatternFill("solid", fgColor="F5F5F5"),
        "secondary": PatternFill("solid", fgColor="E3F2FD"),
    }

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _) in enumerate(HEADERS, start=1):
            val = row.get(key, "") or ""
            cell = ws.cell(r_idx, c_idx, val)
            cell.alignment = wrap
            cell.border = thin
            if key == "source_url" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            q = row.get("data_quality")
            if q in quality_fill:
                cell.fill = quality_fill[q]

    widths = [10, 12, 36, 28, 28, 32, 16, 22, 18, 28, 16, 36, 42, 32, 14, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 28

    meta_ws = wb.create_sheet("说明", 0)
    meta_ws["A1"] = meta.get("title", "NBFC统计")
    meta_ws["A1"].font = Font(bold=True, size=14)
    meta_ws["A3"] = "更新时间"
    meta_ws["B3"] = meta.get("updated", "")
    meta_ws["A4"] = "口径说明"
    meta_ws["B4"] = meta.get("note", "")
    meta_ws["B4"].alignment = wrap
    meta_ws["A5"] = "USD换算"
    meta_ws["B5"] = meta.get("fx_note", "")
    meta_ws["B5"].alignment = wrap
    meta_ws["A7"] = "数据质量"
    meta_ws["B7"] = "official=监管/央行原表；semi-official=协会或监管转引；not_found=仅定位监管机构/类别，数字待补"
    meta_ws["A9"] = "行数"
    meta_ws["B9"] = len(rows)
    meta_ws["A11"] = "JSON源"
    meta_ws["B11"] = "web/src/data/nbfc-country-stats.json"
    meta_ws.column_dimensions["A"].width = 14
    meta_ws.column_dimensions["B"].width = 90

    for out in OUT_PATHS:
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"Wrote {out} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
